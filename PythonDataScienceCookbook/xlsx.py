#coding:utf-8

import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

dest_filename = '/home/coohua/tyh/ltv/data/ltv.xlsx'

ws1 = wb.active
ws1.title = "third_ltv"

# ws1.append('业务月份,注册月份,渠道,arpu,人均成本,生命周期,人均获取成本,ltv'.split(','))


with open(r'/home/coohua/tyh/ltv/data/ltv.csv') as somefile:
    for line in somefile:
        ws1.append(line.split(','))

ws2 = wb.create_sheet(title="second_ltv")

with open(r'/home/coohua/tyh/ltv/data2/ltv.csv') as somefile:
    for line in somefile:
        ws2.append(line.split(','))

ws3 = wb.create_sheet(title="ad_type_income")

with open(r'/home/coohua/tyh/ltv/data3/ad_type.csv') as somefile:
    for line in somefile:
        ws3.append(line.split(','))

wb.save(filename = dest_filename)


